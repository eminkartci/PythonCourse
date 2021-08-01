from array import *
# import openpyxl module
import openpyxl
from openpyxl import Workbook

class Matrix:
    
    #numberOfRows = 0
    #numberOfColumns = 0
    #values = [5][5]p'p

    # A constructor gets again matrix's attributes as parameters, however this time it also takes vectorvalues which was
    # given already in the main method is prepared for the matrix's values. 
    # If vectorvalues length is same as our matrix's dimension, it will take it as values.
    # Otherwise, values of the matrix remains as zero, null.

    def __init__(self, title,NumberOfRows, NumberOfColumns, vectorValues=None):

        self.numberOfRows = NumberOfRows
        self.numberOfColumns = NumberOfColumns
        self.title = title
        if( vectorValues==None):
            self.values = [ [ 0 for y in range( self.numberOfColumns ) ] for x in range( self.numberOfRows ) ]
        elif(len(vectorValues) != NumberOfColumns * NumberOfRows):
            print('the vectorValues array is not suitable for the matrix dimension.')
            self.values = [ [ 0 for y in range( self.numberOfColumns ) ] for x in range( self.numberOfRows ) ]
        else:
            self.values = [ [ 0 for y in range( self.numberOfColumns ) ] for x in range( self.numberOfRows ) ]

            for i in range (self.numberOfRows):
                for j in range(self.numberOfColumns):
                    self.values[i][j] = vectorValues[i *self.numberOfColumns + j]

    # sum up two matrixes which are defined in main method.
    def sum(self, matrix):
        if(self.numberOfRows == matrix.numberOfRows and self.numberOfColumns == matrix.numberOfColumns):
            resultMatrix = Matrix("Result Matrix",self.numberOfRows, self.numberOfColumns)

            for i in range (self.numberOfRows):
                for j in range(self.numberOfColumns):
                    resultMatrix.values[i][j] = self.values[i][j] + matrix.values[i][j]
            
            return resultMatrix
        else:
            return None
    
    #report the result matrix object to the console
    def report(self):
        
        for i in range(self.numberOfRows):
            print('|',end="")
            for j in range(self.numberOfColumns):
                print(self.values[i][j],end="")
                if(j < self.numberOfColumns - 1):
                    print(',',end="")
            print('|')

    def __privateReport(self):
        
        for i in range(self.numberOfRows):
            print('|',end="")
            for j in range(self.numberOfColumns):
                print(self.values[i][j],end="")
                if(j < self.numberOfColumns - 1):
                    print(',',end="")
            print('|')

    def saveMatrixAsExcel(self):
        # Create a Workbook
        wb = Workbook()

        ws =  wb.active
        ws.title = "Matrix " + self.title

        for i in range(self.numberOfColumns):
            for j in range(self.numberOfRows):
                ws.cell(j+1,i+1).value = self.values[j][i]
    

        wb.save(filename = self.title+'.xlsx')



    @staticmethod
    def createMatrixbyExcel(path,title):
        # print("Path:" ,path)
        # workbook object is created
        wb_obj = openpyxl.load_workbook(filename=path,data_only=True)
        ws = wb_obj.active

        # print("Col Num:",ws.max_column)
        # print("Row Num:",ws.max_row)
        values = []

        for i in range(ws.max_column):
            for j in range(ws.max_row):
                values.append(ws.cell(j+1,i+1).value)

        #print("Values:",values)

        tempMatrix = Matrix(title,ws.max_row,ws.max_column,values)
        return tempMatrix

m1 = Matrix.createMatrixbyExcel("matrix1.xlsx","Matrix 1 Baslik")
m2 = Matrix.createMatrixbyExcel("matrix2.xlsx","Matrix 2 Baslik")

mSum = m1.sum(m2)
mSum.report()
mSum.saveMatrixAsExcel()

trialMatrix = Matrix("Finally...",6,2,[2, 4, 3, 52, 3, 4, 2, 423, 4, 32, 3, 242])
trialMatrix.saveMatrixAsExcel()